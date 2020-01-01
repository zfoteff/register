import os
import datetime
import produce
import PySimpleGUI as sg
from openpyxl import load_workbook

################################################################################
#   Script uses PySimpleGUI, openpyxl, as well as other modules
#   to create a cash register system that is based on the Thompson Farms
#   2019 pricing. The program will calculate total cost, display itemized
#   receipt, and updates a separate spreadsheet with a record of the transaction
#   Each entry into the spreadsheet will be a new row on the spreadsheet with:
#       - a timestamp of the time the transaction was entered in the GUI
#       - a record of the total cost
#       - a record of the items sold
################################################################################

# Workbook
wb = load_workbook('TransactionSheet.xlsx')
sheet = wb['Sheet1']

# Dynamic array to store produce items in the receipt & variable for the total price
# that is updated throught the program
receipt = []
total_price = 0.0

# Layout of GUI
button_size = [14, 5]
receipt_window = [[sg.Multiline(text_color='blue', autoscroll=True, size=[80, 20], key='rcpt_window')]]

gui_row1 = [sg.Button('Leeks', key='leek_button', size=button_size),
            sg.Button('Rhubarb', key="rhub_button", size=button_size),
            sg.Button('Asparagus', key='aspa_button', size=button_size), sg.VerticalSeparator(),
            sg.Text(("Total Price: $ "+str(total_price)), auto_size_text=True)]

gui_row2 = [sg.Button('Cauliflower', key='caul_button', size=button_size),
            sg.Button('Broccoli', key="broc_button", size=button_size),
            sg.Button('Carrots', key='crrt_button', size=button_size), sg.VerticalSeparator(),
            sg.Checkbox('+ 1', default=False, key='plusone')]

gui_row3 = [sg.Button('Kale', key='kale_button', size=button_size),
            sg.Button('Spinach', key="spin_button", size=button_size),
            sg.Button('Lettuce', key='lett_button', size=button_size), sg.VerticalSeparator(),
            sg.Checkbox('+ 5', default=False,  key='plusfive')]

gui_row4 = [sg.Button('Beets', key='beet_button', size=button_size),
            sg.Button('Potatoes', key="pota_button", size=button_size),
            sg.Button('Walla-Walla', key='wawa_button', size=button_size), sg.VerticalSeparator(),
            sg.Button('Remove Item', key='remove', size=button_size)]

gui_row5 = [sg.Button('Sweet Onion', key='swon_button', size=button_size),
            sg.Button('Red Onion', key="rdon_button", size=button_size),
            sg.Button('Green Onions', key='gron_button', size=button_size), sg.VerticalSeparator(),
            sg.Button('Checkout', key='check', size=button_size)]

gui_row6 = [sg.Button('Strawberry', key='strw_button'),
            sg.Checkbox('1 pint', default=False, key='pint'),
            sg.Checkbox('4 pack', default=False, key='fourpack'),
            sg.Checkbox('6 pack', default=False, key='sixpack'),
            sg.Checkbox('12 pack', default=False, key='crate')]

gui_hozsep = [sg.Text('________________________________________________________')]

layout = [
          [sg.Column(receipt_window)],
          gui_hozsep,
          gui_row1,
          gui_row2,
          gui_row3,
          gui_row4,
          gui_row5,
          gui_hozsep,
          gui_row6
         ]

window = sg.Window('Register', layout)



################################################################################
# Pre:  an instance of wb exists and has at least 1 open cell
# Post: 3 cells are filled with a time stamp, list of items, and total balance
#       then the document is saved
################################################################################
def update_sheet(transaction_amount, items):
    curtime = datetime.datetime.now()
    timestamp_col = sheet['A']
    transaction_col = sheet['B']
    items_col = sheet['C']

    for cells in timestamp_col:
        if cells.value == None:    # finds first empty cell in timestamp column
            sheet.cell(cells.row, cells.column).value = curtime

    for cells in transaction_col:
        if cells.value == None:    # finds first empty cell in transaction column
            sheet.cell(cells.row, cells.column).value = "$ "+str(transaction_amount)

    for cells in items_col:
        sheet.cell(cells.row, cells.column).value = str(items)

    wb.save(os.path.join('TransactionSheet.xlsx'))


################################################################################
# Pre:  an instance of receipt array exists and has >= 1 element
# Post: Returns an array of strings of items and prices to be displayed in
#       multiline element
#       discount at Thompson Farms
################################################################################
def update_reciept(rcpt_arr):
    string_array = []
    num_discountable_items = 0

    for each in rcpt_arr:
        if each.discount_legal is True:
            num_discountable_items += 1

    if num_discountable_items % 2 != 0:
        num_discountable_items -= 1

    for item in rcpt_arr:
        if item.discount_legal is True and num_discountable_items > 0:
            string_array.append(item.display_transaction(item, True))

        else:
            string_array.append(item.display_transaction(item, False))

    return string_array

################################################################################
# Pre:  an instance of str_arr exists and has > 1 element
# Post  multiline element of window is updated with purchased items and prices
################################################################################
def update_screen(str_arr, multi_element):
    pass


################################################################################
# Pre:  an instance of receipt dictionary exists and has >= 1 element
# Post: the receipt window element is updated with the correct item(s) and prices
#       by referencing the receipt dictionary
################################################################################
def add_to_receipt(rcpt_arr, item, amount_sold):
    counter = 0
    while counter < amount_sold:
        rcpt_arr.append(item)
        counter += 1


# Event loop
while True:
    event, value = window.Read()
    item_amount = 0     # amount of items to be added to the receipt. dependant on
                        # user choice

    if window.Element('plusone').Get():
        item_amount = 1
    elif window.Element('plusfive').Get():
        item_amount = 5
    else:
        item_amount = 1

    if event == 'check':
        update_sheet(total_price, receipt)
        total_price = 0
        receipt.clear()

    if event == 'remove':
        pass

    if event == 'Quit' or event is None:
        update_sheet(total_price, receipt)
        window.Close()
        break

# item buttons
    if event == 'strw_button':  # updates receipt with amount of strawberries
        if window.Element('pint').Get():
            add_to_receipt(receipt, 'strw_pint', 1)

        elif window.Element('fourpack').Get():
            add_to_receipt(receipt, 'strw_four', 1)

        elif window.Element('sixpack').Get():
            add_to_receipt(receipt, 'strw_six', 1)

        elif window.Element('crate').Get():
            add_to_receipt(receipt, 'strw_pack', 1)

        else:
            pass


    if event == 'leek_button':
        add_to_receipt(receipt, produce.Leek, item_amount)
        rcpt_string = str(update_reciept(receipt))
        window.Element('rcpt_window').Update(rcpt_string)

    if event == 'rhub_button':
        add_to_receipt(receipt, produce.Rhubarb, item_amount)
        update_reciept(receipt)

    if event == 'aspa_button':
        add_to_receipt(receipt, produce.Asparagus, item_amount)
        update_reciept(receipt)

    if event == 'caul_button':
        add_to_receipt(receipt, produce.Cauliflower, item_amount)
        update_reciept(receipt)

    if event == 'broc_button':
        add_to_receipt(receipt, produce.Broccoli, item_amount)
        update_reciept(receipt)

    if event == 'crrt_button':
        add_to_receipt(receipt, produce.Carrot, item_amount)
        update_reciept(receipt)

    if event == 'kale_button':
        add_to_receipt(receipt, produce.Kale, item_amount)
        update_reciept(receipt)

    if event == 'spin_button':
        add_to_receipt(receipt, produce.Spinach, item_amount)
        update_reciept(receipt)

    if event == 'lett_button':
        add_to_receipt(receipt, produce.Lettuce, item_amount)
        update_reciept(receipt)

    if event == 'beet_button':
        add_to_receipt(receipt, produce.Beet, item_amount)
        update_reciept(receipt)

    if event == 'pota_button':
        add_to_receipt(receipt, produce.Tater, item_amount)
        update_reciept(receipt)

    if event == 'wawa_button':
        add_to_receipt(receipt, produce.Walla_Walla, item_amount)
        update_reciept(receipt)

    if event == 'swon_button':
        add_to_receipt(receipt, produce.Sweet_Onion, item_amount)
        update_reciept(receipt)

    if event == 'rdon_button':
        add_to_receipt(receipt, produce.Red_Onion, item_amount)
        update_reciept(receipt)

    if event == 'gron_button':
        add_to_receipt(receipt, produce.Green_Onion, item_amount)
        rcpt_string = str(update_reciept(receipt))
        window.Element('rcpt_window').Update(rcpt_string)
